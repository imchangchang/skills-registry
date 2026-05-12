#!/usr/bin/env python3
"""
build_xlsx.py — 把「多专家分析结果 JSON + 原数据 xlsx/csv」渲染为温暖专业风格的 Excel 报告。

视觉哲学（跟HTML报告保持一致）：
- 温暖专业：三文鱼粉封面色、深棕灰正文、非纯黑
- 10米可读：标题字号大、表格有斑马纹、关键数据加粗
- 数据诚实：条件格式只给真实异常上色，不制造视觉噪音

产出：
    Sheet 1 「核心摘要」    结论先行的看板（big numbers + 核心结论段落）
    Sheet 2 「关键指标」    各专家提炼的关键指标合并表（条件格式高亮）
    Sheet 3 「透视_结构」  可选——若原数据适合做类别×度量透视，自动生成
    Sheet 4+ 「原始_xx」    原数据每张表一份（带筛选、冻结首行、斑马纹）

用法：
    python3 build_xlsx.py \\
        --data finance.xlsx \\
        --analysis analysis.json \\
        --output report.xlsx

analysis.json 的约定结构（与 multi-expert-workflow.md 输出 schema 一致）：
    {
      "title": "2026 H1 财务体检报告",
      "summary": "三句话以内的最关键结论",
      "headline_numbers": [
        {"label": "总营收", "value": "1,240 万", "delta": "+12% YoY", "tone": "good"}
      ],
      "experts": [
        {
          "name": "趋势分析师",
          "findings": ["..."],
          "metrics": [{"指标": "月营收增速", "值": "8.2%", "判断": "稳健"}]
        }
      ],
      "pivot_hint": {
        "sheet": "订单", "index": "部门", "columns": "月份",
        "values": "金额", "aggfunc": "sum"
      }
    }

依赖：openpyxl >= 3.1, pandas >= 2.0
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path


def ensure_dependencies() -> None:
    missing = []
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")
    try:
        import pandas  # noqa: F401
    except ImportError:
        missing.append("pandas")
    if missing:
        print(f"[build_xlsx] 缺少依赖: {', '.join(missing)}", file=sys.stderr)
        print(f"[build_xlsx] 请运行: pip install {' '.join(missing)}", file=sys.stderr)
        sys.exit(1)


ensure_dependencies()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# ---------- 温暖专业调色板（对齐 Financial Times 报告风格） ----------

COLOR_BG_WARM = "FFF1E5"         # 三文鱼粉
COLOR_CARD = "FFFFFF"
COLOR_TEXT_MAIN = "33302E"       # 深棕灰
COLOR_TEXT_MUTED = "6B6B6B"
COLOR_ACCENT = "0F5499"          # FT 蓝
COLOR_GOOD = "09823A"
COLOR_WARN = "B8860B"
COLOR_BAD = "CC0000"
COLOR_BORDER = "E0D3C3"
COLOR_STRIPE = "FAF3E8"

THIN = Side(style="thin", color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_TITLE = Font(name="PingFang SC", size=22, bold=True, color=COLOR_TEXT_MAIN)
FONT_SUBTITLE = Font(name="PingFang SC", size=13, color=COLOR_TEXT_MUTED)
FONT_H2 = Font(name="PingFang SC", size=14, bold=True, color=COLOR_ACCENT)
FONT_BODY = Font(name="PingFang SC", size=11, color=COLOR_TEXT_MAIN)
FONT_MUTED = Font(name="PingFang SC", size=10, color=COLOR_TEXT_MUTED)
FONT_BIG_NUMBER = Font(name="PingFang SC", size=20, bold=True, color=COLOR_TEXT_MAIN)
FONT_HEADER = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")

FILL_WARM = PatternFill("solid", fgColor=COLOR_BG_WARM)
FILL_CARD = PatternFill("solid", fgColor=COLOR_CARD)
FILL_HEADER = PatternFill("solid", fgColor=COLOR_TEXT_MAIN)
FILL_STRIPE = PatternFill("solid", fgColor=COLOR_STRIPE)


@dataclass
class Analysis:
    title: str
    summary: str
    headline_numbers: list
    experts: list
    pivot_hint: dict | None

    @staticmethod
    def empty(default_title: str) -> "Analysis":
        return Analysis(title=default_title, summary="", headline_numbers=[],
                        experts=[], pivot_hint=None)

    @staticmethod
    def load(path, fallback_title: str) -> "Analysis":
        if not path or not Path(path).exists():
            return Analysis.empty(fallback_title)
        try:
            raw = json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception as e:
            print(f"[build_xlsx] 读取分析JSON失败，改用空壳: {e}", file=sys.stderr)
            return Analysis.empty(fallback_title)
        return Analysis(
            title=raw.get("title") or fallback_title,
            summary=raw.get("summary") or "",
            headline_numbers=list(raw.get("headline_numbers") or []),
            experts=list(raw.get("experts") or []),
            pivot_hint=raw.get("pivot_hint"),
        )


def load_sheets(data_path: Path) -> dict:
    """返回 {sheet_name: DataFrame}；csv 按单表名返回。"""
    suffix = data_path.suffix.lower()
    if suffix in (".xlsx", ".xls", ".xlsm"):
        return pd.read_excel(data_path, sheet_name=None)
    if suffix == ".csv":
        for enc in ("utf-8", "utf-8-sig", "gbk"):
            try:
                df = pd.read_csv(data_path, encoding=enc)
                return {data_path.stem: df}
            except UnicodeDecodeError:
                continue
        raise RuntimeError(f"CSV 编码识别失败: {data_path}")
    raise RuntimeError(f"暂不支持的数据格式: {suffix}")


def set_row_height(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


def set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def merge_and_write(ws, row: int, col_start: int, col_end: int, value,
                    font: Font, fill=None, align: str = "left") -> None:
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)


def tone_to_font(tone: str) -> Font:
    color = {
        "good": COLOR_GOOD, "warn": COLOR_WARN, "bad": COLOR_BAD,
    }.get((tone or "").lower(), COLOR_TEXT_MAIN)
    return Font(name="PingFang SC", size=12, bold=True, color=color)


def _is_numeric_series(s) -> bool:
    try:
        return pd.api.types.is_numeric_dtype(s)
    except Exception:
        return False


def render_summary_sheet(wb: Workbook, analysis: Analysis, generated_at: str) -> None:
    ws = wb.active
    ws.title = "核心摘要"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, [20, 20, 20, 20, 20, 20])

    # 全幅暖色底
    for r in range(1, 40):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = FILL_WARM

    # 蓝色顶线
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    set_row_height(ws, 1, 6)

    set_row_height(ws, 3, 38)
    merge_and_write(ws, 3, 1, 6, analysis.title or "数据分析报告", FONT_TITLE)

    set_row_height(ws, 4, 22)
    merge_and_write(ws, 4, 1, 6, f"生成于 {generated_at}", FONT_SUBTITLE)

    set_row_height(ws, 6, 20)
    merge_and_write(ws, 6, 1, 6, "核心结论", FONT_H2)

    summary_text = analysis.summary or "（无分析摘要。传入 analysis.json 可渲染结论段落。）"
    set_row_height(ws, 7, max(60, 18 * (summary_text.count("\n") + 2)))
    merge_and_write(ws, 7, 1, 6, summary_text,
                    Font(name="PingFang SC", size=12, color=COLOR_TEXT_MAIN),
                    fill=FILL_CARD)
    for c in range(1, 7):
        ws.cell(row=7, column=c).border = BORDER_ALL

    set_row_height(ws, 9, 20)
    merge_and_write(ws, 9, 1, 6, "关键数字", FONT_H2)

    numbers = analysis.headline_numbers[:6]
    for idx, item in enumerate(numbers):
        row_block = 10 + (idx // 3) * 4
        col_block = 1 + (idx % 3) * 2
        set_row_height(ws, row_block, 18)
        merge_and_write(ws, row_block, col_block, col_block + 1,
                        item.get("label", ""), FONT_MUTED, fill=FILL_CARD)
        set_row_height(ws, row_block + 1, 36)
        merge_and_write(ws, row_block + 1, col_block, col_block + 1,
                        item.get("value", ""), FONT_BIG_NUMBER, fill=FILL_CARD)
        set_row_height(ws, row_block + 2, 18)
        merge_and_write(ws, row_block + 2, col_block, col_block + 1,
                        item.get("delta", ""), tone_to_font(item.get("tone", "")),
                        fill=FILL_CARD)
        for r in range(row_block, row_block + 3):
            for c in range(col_block, col_block + 2):
                ws.cell(row=r, column=c).border = BORDER_ALL

    foot_row = 22 + ((max(len(numbers), 1) - 1) // 3) * 4
    merge_and_write(ws, foot_row, 1, 6,
                    "© 由 huashu-data-pro 生成，数据诚实，结论先行。",
                    FONT_MUTED)


def render_metrics_sheet(wb: Workbook, analysis: Analysis) -> None:
    if not analysis.experts:
        return
    ws = wb.create_sheet("关键指标")
    ws.sheet_view.showGridLines = False

    headers = ["视角", "指标", "值", "判断"]
    set_col_widths(ws, [14, 32, 18, 40])

    merge_and_write(ws, 1, 1, 4, "各视角提炼的关键指标", FONT_H2, fill=FILL_WARM)
    set_row_height(ws, 1, 28)

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER_ALL
    set_row_height(ws, 3, 24)

    row = 4
    for expert in analysis.experts:
        expert_name = expert.get("name", "未命名视角")
        metrics = expert.get("metrics") or []
        if not metrics:
            findings = expert.get("findings") or []
            metrics = [{"指标": "核心发现", "值": "-", "判断": (f or "")[:80]}
                       for f in findings[:3]]
        for m in metrics:
            values = [expert_name, m.get("指标", ""), m.get("值", ""), m.get("判断", "")]
            for i, v in enumerate(values, start=1):
                c = ws.cell(row=row, column=i, value=v)
                c.font = FONT_BODY
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)
                c.border = BORDER_ALL
                if row % 2 == 0:
                    c.fill = FILL_STRIPE
            set_row_height(ws, row, 22)
            row += 1

    ws.freeze_panes = "A4"


def render_pivot_sheet(wb: Workbook, sheets: dict, pivot_hint) -> None:
    hint = pivot_hint or {}
    sheet_name = hint.get("sheet")
    df = None
    if sheet_name and sheet_name in sheets:
        df = sheets[sheet_name]
    elif sheets:
        for _name, _df in sheets.items():
            if _df.shape[1] >= 3:
                df = _df
                sheet_name = _name
                break
    if df is None or df.empty:
        return

    index = hint.get("index")
    columns = hint.get("columns")
    values = hint.get("values")
    aggfunc = hint.get("aggfunc") or "sum"

    if not values:
        numeric_cols = [c for c in df.columns if _is_numeric_series(df[c])]
        if not numeric_cols:
            return
        values = numeric_cols[0]
    if not index:
        cat_cols = [c for c in df.columns
                    if not _is_numeric_series(df[c]) and c != values]
        if not cat_cols:
            return
        index = cat_cols[0]
    if not columns:
        cat_cols = [c for c in df.columns
                    if not _is_numeric_series(df[c]) and c not in (values, index)]
        columns = cat_cols[0] if cat_cols else None

    try:
        if columns:
            pivot = pd.pivot_table(df, index=index, columns=columns,
                                   values=values, aggfunc=aggfunc, fill_value=0)
        else:
            pivot = df.groupby(index)[values].agg(aggfunc).to_frame(values)
    except Exception as e:
        print(f"[build_xlsx] 透视表生成失败: {e}", file=sys.stderr)
        return

    ws = wb.create_sheet("透视_结构")
    ws.sheet_view.showGridLines = False

    merge_and_write(ws, 1, 1, max(2, pivot.shape[1] + 1),
                    f"透视：{index} × {columns or '—'} → {values}（{aggfunc}）",
                    FONT_H2, fill=FILL_WARM)
    set_row_height(ws, 1, 28)

    ws.cell(row=3, column=1, value=str(index))
    ws.cell(row=3, column=1).font = FONT_HEADER
    ws.cell(row=3, column=1).fill = FILL_HEADER
    ws.cell(row=3, column=1).border = BORDER_ALL
    for j, col_name in enumerate(pivot.columns, start=2):
        c = ws.cell(row=3, column=j, value=str(col_name))
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = BORDER_ALL
    set_row_height(ws, 3, 22)

    for i, (idx_val, row_data) in enumerate(pivot.iterrows(), start=4):
        c = ws.cell(row=i, column=1, value=str(idx_val))
        c.font = FONT_BODY
        c.border = BORDER_ALL
        if i % 2 == 0:
            c.fill = FILL_STRIPE
        for j, val in enumerate(row_data.values, start=2):
            vc = ws.cell(row=i, column=j, value=float(val) if pd.notna(val) else 0)
            vc.font = FONT_BODY
            vc.alignment = Alignment(horizontal="right")
            vc.number_format = "#,##0.00"
            vc.border = BORDER_ALL
            if i % 2 == 0:
                vc.fill = FILL_STRIPE

    last_row = 3 + pivot.shape[0]
    last_col = 1 + pivot.shape[1]
    if pivot.shape[1] > 0 and pivot.shape[0] > 0:
        rng = f"{get_column_letter(2)}4:{get_column_letter(last_col)}{last_row}"
        scale = ColorScaleRule(
            start_type="min", start_color="FFF3E0",
            mid_type="percentile", mid_value=50, mid_color="F6D9B3",
            end_type="max", end_color="E17055",
        )
        ws.conditional_formatting.add(rng, scale)

    widths = [max(14, min(40, len(str(index)) * 2 + 8))]
    for col_name in pivot.columns:
        widths.append(max(12, min(26, len(str(col_name)) * 2 + 6)))
    set_col_widths(ws, widths)
    ws.freeze_panes = "B4"


def render_data_sheets(wb: Workbook, sheets: dict) -> None:
    for name, df in sheets.items():
        if df is None or df.empty:
            continue
        safe = f"原始_{str(name)[:25]}"
        ws = wb.create_sheet(safe)
        ws.sheet_view.showGridLines = False

        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=j, value=str(col))
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = BORDER_ALL
        set_row_height(ws, 1, 24)

        for i, (_, row) in enumerate(df.iterrows(), start=2):
            for j, val in enumerate(row.values, start=1):
                write_val = val
                if pd.isna(val):
                    write_val = None
                elif hasattr(val, "to_pydatetime"):
                    write_val = val.to_pydatetime()
                c = ws.cell(row=i, column=j, value=write_val)
                c.font = FONT_BODY
                c.border = BORDER_ALL
                if i % 2 == 0:
                    c.fill = FILL_STRIPE
                if isinstance(write_val, (int, float)) and not isinstance(write_val, bool):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")

        for j, col in enumerate(df.columns, start=1):
            sample = df[col].astype(str).head(20).tolist() + [str(col)]
            width = min(32, max(10, max(len(s) for s in sample) + 2))
            ws.column_dimensions[get_column_letter(j)].width = width

        ws.freeze_panes = "A2"
        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

        numeric_cols_added = 0
        last_row = 1 + len(df)
        for j, col in enumerate(df.columns, start=1):
            if numeric_cols_added >= 5:
                break
            if _is_numeric_series(df[col]):
                col_letter = get_column_letter(j)
                scale = ColorScaleRule(
                    start_type="min", start_color="F7D7D7",
                    mid_type="percentile", mid_value=50, mid_color="FFF3E0",
                    end_type="max", end_color="D4EDDA",
                )
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{last_row}", scale)
                numeric_cols_added += 1


def build(data_path: Path, analysis_path, output_path: Path,
          report_title=None) -> None:
    from datetime import datetime

    sheets = load_sheets(data_path)
    fallback_title = report_title or f"{data_path.stem} 分析报告"
    analysis = Analysis.load(analysis_path, fallback_title) if analysis_path \
        else Analysis.empty(fallback_title)

    wb = Workbook()
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    render_summary_sheet(wb, analysis, generated_at)
    render_metrics_sheet(wb, analysis)
    render_pivot_sheet(wb, sheets, analysis.pivot_hint)
    render_data_sheets(wb, sheets)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[build_xlsx] 已生成: {output_path}")


def parse_args():
    p = argparse.ArgumentParser(description="渲染温暖专业风格的 Excel 分析报告。")
    p.add_argument("--data", required=True, help="原始数据文件（.xlsx 或 .csv）")
    p.add_argument("--analysis", required=False,
                   help="多专家分析结果 JSON；不给则只渲染数据")
    p.add_argument("--output", required=True, help="输出 .xlsx 路径")
    p.add_argument("--title", required=False, help="报告标题（覆盖 analysis.json.title）")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    data_path = Path(args.data).expanduser().resolve()
    if not data_path.exists():
        print(f"[build_xlsx] 数据文件不存在: {data_path}", file=sys.stderr)
        sys.exit(2)
    analysis_path = Path(args.analysis).expanduser().resolve() if args.analysis else None
    if analysis_path and not analysis_path.exists():
        print(f"[build_xlsx] 分析JSON不存在，将跳过: {analysis_path}", file=sys.stderr)
        analysis_path = None

    output_path = Path(args.output).expanduser().resolve()
    try:
        build(data_path, analysis_path, output_path, report_title=args.title)
    except Exception as e:
        print(f"[build_xlsx] 构建失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
